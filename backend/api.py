"""
backend/api.py — Rotas da API REST do Hub de Automações.
"""

import csv
import io
import json
import logging
import os
import threading
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Security
from fastapi.responses import JSONResponse, Response, StreamingResponse
from fastapi.security import APIKeyHeader

logger = logging.getLogger(__name__)

_KEY_HEADER = APIKeyHeader(name="X-Hub-API-Key", auto_error=False)


async def _verificar_key(key: str | None = Security(_KEY_HEADER)) -> None:
    expected = os.getenv("HUB_API_KEY", "")
    if not expected:
        raise HTTPException(
            status_code=503,
            detail="Servidor mal configurado: defina HUB_API_KEY no .env.",
        )
    if key != expected:
        raise HTTPException(status_code=401, detail="API key inválida ou ausente.")


router = APIRouter(dependencies=[Depends(_verificar_key)])

_DB_CONTRATACOES = Path("database/contratacoes.json")
_DB_HISTORICO    = Path("database/historico.json")
_DB_USUARIOS     = Path("database/usuarios_analise.json")

_VAZIO_USUARIOS = json.dumps({
    "ultima_atualizacao": None,
    "inativos-maquinas":  [],
    "cc-divergente":      [],
    "multi-responsaveis": [],
}, ensure_ascii=False)

_VAZIO = json.dumps(
    {"ultima_atualizacao": None, "total": 0, "chamados": []},
    ensure_ascii=False,
)

_VAZIO_HIST = json.dumps([], ensure_ascii=False)


def _verdanadesk_base() -> str:
    api_url = os.getenv("API_URL", "")
    idx = api_url.find("/api.php")
    return api_url[:idx] if idx != -1 else api_url.rstrip("/")


@router.get("/config", summary="Configurações públicas do servidor")
async def get_config() -> JSONResponse:
    return JSONResponse({"verdanadesk_url": _verdanadesk_base()})


@router.get("/contratacoes", summary="Chamados de contratação ativos")
async def get_contratacoes() -> Response:
    if not _DB_CONTRATACOES.exists():
        return Response(content=_VAZIO, media_type="application/json")
    try:
        return Response(
            content=_DB_CONTRATACOES.read_text(encoding="utf-8"),
            media_type="application/json",
        )
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"erro": str(exc), "total": 0, "chamados": []},
        )


@router.get("/historico", summary="KPIs históricos por dia")
async def get_historico() -> Response:
    if not _DB_HISTORICO.exists():
        return Response(content=_VAZIO_HIST, media_type="application/json")
    try:
        return Response(
            content=_DB_HISTORICO.read_text(encoding="utf-8"),
            media_type="application/json",
        )
    except Exception as exc:
        return JSONResponse(status_code=500, content={"erro": str(exc)})


@router.get("/contratacoes/export", summary="Exporta chamados em CSV, XLSX ou PDF")
async def export_contratacoes(
    formato: str = Query("xlsx", pattern="^(csv|xlsx|pdf)$"),
    termo: str | None = Query(None),
) -> StreamingResponse:
    if not _DB_CONTRATACOES.exists():
        return JSONResponse(status_code=404, content={"erro": "Dados ainda não disponíveis."})

    data = json.loads(_DB_CONTRATACOES.read_text(encoding="utf-8"))
    chamados: list[dict] = data.get("chamados", [])

    if termo:
        chamados = [c for c in chamados if c.get("Termo_Status") == termo]

    ts_snapshot = data.get("ultima_atualizacao")
    if formato == "csv":
        return _export_csv(chamados, ts_snapshot)
    if formato == "pdf":
        return _export_pdf(chamados, ts_snapshot)
    return _export_xlsx(chamados, ts_snapshot)


# ── Constantes compartilhadas ────────────────────────────────────────────────

_COLUNAS = ["ID_do_Chamado", "Titulo", "Status", "Tempo_Solucao",
            "Data_Abertura", "Requerente", "Termo_Status"]

_HEADERS_PT = {
    "ID_do_Chamado": "ID",
    "Titulo":        "Título",
    "Status":        "Status",
    "Tempo_Solucao": "SLA / Tempo p/ Solução",
    "Data_Abertura": "Data de Abertura",
    "Requerente":    "Requerente",
    "Termo_Status":  "Status do Termo",
}

_STATUS_COLORS = {
    "Novo":           {"fill": "DCFCE7", "font": "166534"},
    "Em Atendimento": {"fill": "DBEAFE", "font": "1D4ED8"},
    "Pendente":       {"fill": "FEF9C3", "font": "854D0E"},
    "Solucionado":    {"fill": "F1F5F9", "font": "475569"},
    "Fechado":        {"fill": "F1F5F9", "font": "475569"},
}

_TERMO_COLORS = {
    "Termo OK":          {"fill": "D1FAE5", "font": "065F46"},
    "Pendente":          {"fill": "FEF3C7", "font": "92400E"},
    "Sem tarefa":        {"fill": "FEE2E2", "font": "991B1B"},
    "Erro ao verificar": {"fill": "FEE2E2", "font": "991B1B"},
}


def _ts_filename(ultima_atualizacao: str | None) -> str:
    try:
        d = datetime.fromisoformat(ultima_atualizacao or "")
        return d.strftime("%Y%m%d_%H%M")
    except Exception:
        return datetime.now().strftime("%Y%m%d_%H%M")


def _fmt_snapshot(ultima_atualizacao: str | None) -> str:
    try:
        d = datetime.fromisoformat(ultima_atualizacao or "")
        return d.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return datetime.now().strftime("%d/%m/%Y %H:%M")


# ── CSV ──────────────────────────────────────────────────────────────────────

def _export_csv(chamados: list[dict], ultima_atualizacao: str | None) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=_COLUNAS,
        extrasaction="ignore",
        lineterminator="\r\n",
    )
    writer.writerow({c: _HEADERS_PT[c] for c in _COLUNAS})
    writer.writerows(chamados)
    ts = _ts_filename(ultima_atualizacao)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="contratacoes_{ts}.csv"'},
    )


# ── XLSX ─────────────────────────────────────────────────────────────────────

def _export_xlsx(chamados: list[dict], ultima_atualizacao: str | None) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse(
            status_code=500,
            content={"erro": "openpyxl não instalado. Execute: pip install openpyxl"},
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratações"

    # ── Paleta ────────────────────────────────────────────────────────────────
    BRAND_DARK   = "0D0D0F"   # fundo do sistema
    BRAND_ORANGE = "FF7A1A"   # accent
    BRAND_NAVY   = "1E293B"   # texto principal

    THIN   = Side(style="thin",   color="D1D5DB")
    MEDIUM = Side(style="medium", color="CBD5E1")
    BORDER_DATA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BORDER_HDR  = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

    def _fill(hex_: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_)

    def _font(hex_: str, size: int = 10, bold: bool = False, italic: bool = False) -> Font:
        return Font(color=hex_, size=size, bold=bold, italic=italic)

    def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Linha 1: Banner do relatório ─────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    banner = ws.cell(row=1, column=1,
                     value="CHAMADOS DE CONTRATAÇÃO ATIVOS  ·  Hub de Automações TI")
    banner.fill      = _fill(BRAND_DARK)
    banner.font      = Font(color=BRAND_ORANGE, size=13, bold=True,
                            name="Calibri")
    banner.alignment = _align("left", "center")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(_COLUNAS))

    # ── Linha 2: Metadados ───────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    snap_str = _fmt_snapshot(ultima_atualizacao)
    meta = ws.cell(
        row=2, column=1,
        value=(f"Snapshot GLPI: {snap_str}    "
               f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}    "
               f"Total: {len(chamados)} chamado(s)"),
    )
    meta.fill      = _fill("1E293B")
    meta.font      = _font("94A3B8", size=9, italic=True)
    meta.alignment = _align("left", "center")
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=len(_COLUNAS))

    # ── Linha 3: Cabeçalhos das colunas ─────────────────────────────────────
    ws.row_dimensions[3].height = 30
    for col_idx, col_key in enumerate(_COLUNAS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=_HEADERS_PT[col_key])
        cell.fill      = _fill("334155")
        cell.font      = Font(bold=True, color="F1F5F9", size=10, name="Calibri")
        cell.alignment = _align("center", "center")
        cell.border    = BORDER_HDR

    # ── Linhas de dados ──────────────────────────────────────────────────────
    DATA_START = 4
    for row_idx, chamado in enumerate(chamados, start=DATA_START):
        ws.row_dimensions[row_idx].height = 21
        is_even = (row_idx % 2 == 0)
        base_fill_hex = "F8FAFC" if is_even else "FFFFFF"

        for col_idx, col_key in enumerate(_COLUNAS, start=1):
            val = chamado.get(col_key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER_DATA

            if col_key == "ID_do_Chamado":
                cell.fill      = _fill(base_fill_hex)
                cell.font      = Font(color="6366F1", size=10, bold=True, name="Calibri")
                cell.alignment = _align("center")

            elif col_key == "Titulo":
                cell.fill      = _fill(base_fill_hex)
                cell.font      = _font(BRAND_NAVY, size=10)
                cell.alignment = _align("left", wrap=True)

            elif col_key == "Status":
                sc = _STATUS_COLORS.get(val, {"fill": base_fill_hex, "font": "64748B"})
                cell.fill      = _fill(sc["fill"])
                cell.font      = Font(color=sc["font"], size=10, bold=True, name="Calibri")
                cell.alignment = _align("center")

            elif col_key == "Tempo_Solucao":
                cell.fill = _fill(base_fill_hex)
                v_lower = str(val).lower()
                if "atraso" in v_lower or "expir" in v_lower:
                    fcolor = "DC2626"
                elif any(x in v_lower for x in ["h ", "hora"]) and "dia" not in v_lower:
                    fcolor = "D97706"
                else:
                    fcolor = "374151"
                cell.font      = Font(color=fcolor, size=10, name="Calibri")
                cell.alignment = _align("center")

            elif col_key == "Termo_Status":
                tc = _TERMO_COLORS.get(val, {"fill": base_fill_hex, "font": "64748B"})
                cell.fill      = _fill(tc["fill"])
                cell.font      = Font(color=tc["font"], size=10, bold=True, name="Calibri")
                cell.alignment = _align("center")

            else:
                cell.fill      = _fill(base_fill_hex)
                cell.font      = _font("475569", size=10)
                cell.alignment = _align("left")

    # ── Larguras automáticas ─────────────────────────────────────────────────
    MIN_W = {"ID_do_Chamado": 7,  "Status": 17, "Termo_Status": 20,
             "Data_Abertura": 19, "Tempo_Solucao": 24}
    MAX_W = {"Titulo": 55, "Requerente": 32}
    for col_idx, col_key in enumerate(_COLUNAS, start=1):
        col_letter = get_column_letter(col_idx)
        all_vals = [_HEADERS_PT[col_key]] + [str(c.get(col_key, "") or "") for c in chamados]
        best_w = max((len(v) for v in all_vals), default=10) + 3
        best_w = max(best_w, MIN_W.get(col_key, 13))
        best_w = min(best_w, MAX_W.get(col_key, 60))
        ws.column_dimensions[col_letter].width = best_w

    # ── Freeze pane + auto-filter ─────────────────────────────────────────────
    ws.freeze_panes = f"A{DATA_START}"
    last_col = get_column_letter(len(_COLUNAS))
    last_row = DATA_START - 1 + max(len(chamados), 1)
    ws.auto_filter.ref = f"A3:{last_col}3"

    # ── Configurações de impressão ────────────────────────────────────────────
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9   # A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_title_rows = "3:3"     # repeat header row when printing

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = _ts_filename(ultima_atualizacao)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="contratacoes_{ts}.xlsx"'},
    )


# ── PDF ──────────────────────────────────────────────────────────────────────

def _export_pdf(chamados: list[dict], ultima_atualizacao: str | None) -> StreamingResponse:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                        Table, TableStyle)
    except ImportError:
        return JSONResponse(
            status_code=500,
            content={"erro": "reportlab não instalado. Execute: pip install reportlab"},
        )

    buf = io.BytesIO()
    PAGE = landscape(A4)
    doc = SimpleDocTemplate(
        buf,
        pagesize=PAGE,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )

    # ── Paleta ────────────────────────────────────────────────────────────────
    C_BRAND  = colors.HexColor("#FF7A1A")
    C_DARK   = colors.HexColor("#0D0D0F")
    C_NAVY   = colors.HexColor("#1E293B")
    C_SLATE  = colors.HexColor("#334155")
    C_HDR_TXT = colors.HexColor("#F1F5F9")
    C_ROW_A  = colors.HexColor("#F8FAFC")
    C_ROW_B  = colors.white
    C_LINE   = colors.HexColor("#E2E8F0")
    C_MUTED  = colors.HexColor("#94A3B8")

    C_OK_BG  = colors.HexColor("#D1FAE5"); C_OK_FG  = colors.HexColor("#065F46")
    C_WN_BG  = colors.HexColor("#FEF3C7"); C_WN_FG  = colors.HexColor("#92400E")
    C_ER_BG  = colors.HexColor("#FEE2E2"); C_ER_FG  = colors.HexColor("#991B1B")

    C_ST_NOVO = {"bg": colors.HexColor("#DCFCE7"), "fg": colors.HexColor("#166534")}
    C_ST_ATND = {"bg": colors.HexColor("#DBEAFE"), "fg": colors.HexColor("#1D4ED8")}
    C_ST_PEND = {"bg": colors.HexColor("#FEF9C3"), "fg": colors.HexColor("#854D0E")}
    C_ST_OTHR = {"bg": colors.HexColor("#F1F5F9"), "fg": colors.HexColor("#475569")}

    STATUS_PAL = {
        "Novo":           C_ST_NOVO,
        "Em Atendimento": C_ST_ATND,
        "Pendente":       C_ST_PEND,
    }
    TERMO_PAL = {
        "Termo OK":          {"bg": C_OK_BG, "fg": C_OK_FG},
        "Pendente":          {"bg": C_WN_BG, "fg": C_WN_FG},
        "Sem tarefa":        {"bg": C_ER_BG, "fg": C_ER_FG},
        "Erro ao verificar": {"bg": C_ER_BG, "fg": C_ER_FG},
    }

    # ── Estilos de parágrafo ─────────────────────────────────────────────────
    def _pstyle(name, font="Helvetica", size=8, color=C_NAVY,
                align=TA_LEFT, bold=False, leading=None):
        return ParagraphStyle(
            name,
            fontName=f"Helvetica-Bold" if bold else font,
            fontSize=size,
            textColor=color,
            alignment=align,
            leading=leading or size * 1.3,
            spaceAfter=0,
            spaceBefore=0,
        )

    P_TITLE   = _pstyle("title",  size=14, color=C_BRAND,  bold=True)
    P_SUB     = _pstyle("sub",    size=8,  color=C_MUTED)
    P_HDR     = _pstyle("hdr",    size=8,  color=C_HDR_TXT, align=TA_CENTER, bold=True)
    P_CELL    = _pstyle("cell",   size=8,  color=C_NAVY,    align=TA_LEFT)
    P_CENTER  = _pstyle("center", size=8,  color=C_NAVY,    align=TA_CENTER)
    P_ID      = _pstyle("id",     size=8,  color=colors.HexColor("#6366F1"),
                         align=TA_CENTER, bold=True)
    P_TITULO  = _pstyle("titulo", size=7.5, color=C_NAVY,   align=TA_LEFT,  leading=10)

    def _status_style(val: str):
        pal = STATUS_PAL.get(val, C_ST_OTHR)
        return _pstyle(f"st_{val}", size=8, color=pal["fg"], align=TA_CENTER, bold=True)

    def _termo_style(val: str):
        pal = TERMO_PAL.get(val, {"bg": C_ROW_A, "fg": C_SLATE})
        return _pstyle(f"tr_{val}", size=8, color=pal["fg"], align=TA_CENTER, bold=True)

    # ── Cabeçalho do documento ────────────────────────────────────────────────
    snap_str = _fmt_snapshot(ultima_atualizacao)
    now_str  = datetime.now().strftime("%d/%m/%Y %H:%M")
    story = [
        Paragraph("CHAMADOS DE CONTRATAÇÃO ATIVOS", P_TITLE),
        Spacer(1, 0.15 * cm),
        Paragraph(
            f"Hub de Automações TI  ·  Snapshot GLPI: {snap_str}  ·  "
            f"Exportado em: {now_str}  ·  Total: {len(chamados)} chamado(s)",
            P_SUB,
        ),
        Spacer(1, 0.4 * cm),
    ]

    # ── Tabela ────────────────────────────────────────────────────────────────
    # Larguras das colunas (total ≈ página landscape A4 - margens)
    PAGE_W = PAGE[0] - 3 * cm
    COL_W = {
        "ID_do_Chamado": 1.5  * cm,
        "Titulo":        8.5  * cm,
        "Status":        3.0  * cm,
        "Tempo_Solucao": 4.0  * cm,
        "Data_Abertura": 3.2  * cm,
        "Requerente":    4.5  * cm,
        "Termo_Status":  3.8  * cm,
    }
    col_widths = [COL_W[c] for c in _COLUNAS]

    # Linha de cabeçalho
    header_row = [Paragraph(_HEADERS_PT[c], P_HDR) for c in _COLUNAS]
    table_data = [header_row]

    # Linhas de dados + estilos condicionais por célula
    cell_styles: list[tuple] = []
    for row_i, chamado in enumerate(chamados):
        data_row_idx = row_i + 1  # 0 = header
        is_even = (row_i % 2 == 0)
        row_bg = C_ROW_A if is_even else C_ROW_B
        row: list = []

        for col_key in _COLUNAS:
            val = str(chamado.get(col_key, "") or "")

            if col_key == "ID_do_Chamado":
                row.append(Paragraph(f"#{val}", P_ID))

            elif col_key == "Titulo":
                row.append(Paragraph(val, P_TITULO))

            elif col_key == "Status":
                pal = STATUS_PAL.get(val, C_ST_OTHR)
                col_idx = _COLUNAS.index(col_key)
                cell_styles.append(
                    ("BACKGROUND", (col_idx, data_row_idx), (col_idx, data_row_idx), pal["bg"])
                )
                row.append(Paragraph(val, _status_style(val)))

            elif col_key == "Tempo_Solucao":
                v_lower = val.lower()
                if "atraso" in v_lower or "expir" in v_lower:
                    fg = colors.HexColor("#DC2626")
                elif any(x in v_lower for x in ["h ", "hora"]) and "dia" not in v_lower:
                    fg = colors.HexColor("#D97706")
                else:
                    fg = C_NAVY
                st = _pstyle(f"sla_{row_i}", size=8, color=fg, align=TA_CENTER)
                row.append(Paragraph(val, st))

            elif col_key == "Termo_Status":
                pal = TERMO_PAL.get(val, {"bg": row_bg, "fg": C_SLATE})
                col_idx = _COLUNAS.index(col_key)
                cell_styles.append(
                    ("BACKGROUND", (col_idx, data_row_idx), (col_idx, data_row_idx), pal["bg"])
                )
                row.append(Paragraph(val, _termo_style(val)))

            else:
                row.append(Paragraph(val, P_CELL))

        table_data.append(row)

    # ── Estilo base da tabela ─────────────────────────────────────────────────
    base_style = [
        # Cabeçalho
        ("BACKGROUND",  (0, 0), (-1, 0),  C_SLATE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_ROW_A, C_ROW_B]),
        # Bordas
        ("GRID",        (0, 0), (-1, -1), 0.4, C_LINE),
        ("LINEABOVE",   (0, 0), (-1, 0),  1.0, C_BRAND),
        ("LINEBELOW",   (0, 0), (-1, 0),  0.8, C_SLATE),
        # Padding
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        # Alinhamento vertical
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
    ]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(base_style + cell_styles))
    story.append(tbl)

    # ── Rodapé (via canvasmaker) ──────────────────────────────────────────────
    def _add_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(C_MUTED)
        w, h = PAGE
        canvas.drawString(1.5 * cm, 0.8 * cm,
                          f"Hub de Automações TI  ·  {now_str}")
        canvas.drawRightString(w - 1.5 * cm, 0.8 * cm,
                               f"Página {doc_.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_add_footer, onLaterPages=_add_footer)
    buf.seek(0)

    ts = _ts_filename(ultima_atualizacao)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="contratacoes_{ts}.pdf"'},
    )


# ── Análise de Usuários ───────────────────────────────────────────────────────

@router.get("/analise-usuarios", summary="Retorna os dados das 3 análises de inventário")
async def get_analise_usuarios() -> Response:
    if not _DB_USUARIOS.exists():
        return Response(content=_VAZIO_USUARIOS, media_type="application/json")
    try:
        return Response(
            content=_DB_USUARIOS.read_text(encoding="utf-8"),
            media_type="application/json",
        )
    except Exception as exc:
        return JSONResponse(status_code=500, content={"erro": str(exc)})


@router.post("/analise-usuarios/sync", summary="Dispara nova varredura de usuários × máquinas")
async def sync_analise_usuarios(background_tasks: BackgroundTasks) -> JSONResponse:
    from automations.usuarios import usuarios_sync

    if usuarios_sync.is_running():
        return JSONResponse(
            status_code=409,
            content={"ok": False, "message": "Sincronização já em andamento."},
        )

    def _run():
        try:
            from automations.contratacoes.glpi_sync import ClienteGLPI, carregar_configuracoes
            config = carregar_configuracoes()
            glpi = ClienteGLPI(
                base_url=config["API_URL"],
                token_url=config["OAUTH_TOKEN_URL"],
                client_id=config["OAUTH_CLIENT_ID"],
                client_secret=config["OAUTH_CLIENT_SECRET"],
                username=config["OAUTH_USERNAME"],
                password=config["OAUTH_PASSWORD"],
                categoria_ids=config["CATEGORIA_IDS"],
                cats_com_ativo=config["CATS_COM_ATIVO"],
            )
            usuarios_sync.executar(glpi)
        except Exception as exc:
            logger.error("Erro na sincronização de usuários: %s", exc)

    background_tasks.add_task(_run)
    return JSONResponse({"ok": True, "message": "Sincronização iniciada em background."})


@router.get("/analise-usuarios/status", summary="Verifica se a sincronização está em andamento")
async def status_sync_usuarios() -> JSONResponse:
    from automations.usuarios import usuarios_sync
    return JSONResponse({"running": usuarios_sync.is_running()})
